import os
import time
from colorama import Fore,init
import psycopg
import sys
import subprocess

from openpyxl.styles import Border, Side

import Utilities
from Config import pc_tests_hide, hide_calibration_params, cold_only
from Utilities import convert_scientific_to_float, SELECT_PHASE
from WorkbookCreator import *
import matplotlib.pyplot as mp
import numpy as np
init(autoreset=True)

#this also acts a terminal to run database management
#updates here too
DB_INFO = {
    "host": "localhost",
    "dbname": "dcdc_tests",
    "user": "studadmin",
    "password": "password",
    "port": 5432
}

#Run this file to update data in your local spreadsheet, but replace the following parameter with system desktop path or
#other desired download destination:
DD = rf"C:\Users\StudentAdmin\Desktop"
XLSX_NAME= "TESTS"

#clutter hiding options
if pc_tests_hide:
    hide_sheets = ["Multiple Power Cycle Cold","Multiple Power Cycle Warm"]
else:
    hide_sheets = []
    if cold_only:
        hide_sheets = ["Multiple Power Cycle Warm"]

def Trace_Getter(board_id, phase, output_path):
    '''Writes the data for the trace of a board to a file'''
    #again im very iffy about sql atm but this should work mostly fine
    with (get_connection() as conn):
        with conn.cursor() as cursor:

            cursor.execute("""
                SELECT *
                FROM board_traces
                WHERE board_id = %s
                  AND phase = %s
            """, (board_id, phase))

            row = cursor.fetchone()

            if row is None:
                sys.exit(f"No trace data found for board {board_id}")

            columns = [desc[0] for desc in cursor.description]
        data = dict(zip(columns, row))

        trace_pairs = {
            "Input Voltage Sweep Warm": (
                "input_voltage_sweep_voltage_trace",
                "input_voltage_sweep_current_trace"
            ),
            "Nominal Load Warm": (
                "nominal_load_voltage_trace",
                "nominal_load_current_trace"
            ),
            "Multiple Power Cycle Warm": (
                "multiple_power_cycle_voltage",
                "multiple_power_cycle_current"
            ),
            "Multiple Power Cycle Cold": (
                "multiple_power_cycle_voltage_c",
                "multiple_power_cycle_current_c"
            ),
            "Input Voltage Step (5 to 5.1)": (
                "input_step_voltage_voltage_trace",
                "input_step_voltage_current_trace"
            ),
            "Nominal Load Cold": (
                "output_step_load_voltage_trace",
                "output_step_load_current_trace"
            ),
            "Input Voltage Sweep Cold": (
                "input_voltage_sweep_voltage_trace_c",
                "input_voltage_sweep_current_trace_c"
            ),
            "Cold Start Up": (
                "cold_startup_voltage_trace",
                "cold_startup_current_trace"
            )
        }

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

            for name, (vol_col, cur_col) in trace_pairs.items():
                #hide clutter sheets if needed
                if name in hide_sheets:
                    continue

                voltage = data.get(vol_col)
                current = data.get(cur_col)

                if voltage is not None:

                    voltage = json.loads(voltage)

                    output = {
                        "Sample": range(len(voltage)),
                        "Voltage": voltage
                    }

                    if current is not None:
                        current = json.loads(current)

                        #add current if it matches voltage length
                        if len(current) == len(voltage):
                            if name == "Input Voltage Step (5 to 5.1)" or name == "Nominal Load Cold":
                                output["Input Voltage"] = current
                            if name == "Multiple Power Cycle Cold":
                                #for this column we care about different results
                                output["Input Voltage"] = current
                            else:
                                output["Current"] = current

                    df = pd.DataFrame(output)

                    df = df.map(convert_scientific_to_float)

                    df.to_excel(
                        writer,
                        sheet_name=name[:31],
                        index=False
                    )

   #scientific notation
                    ws = writer.sheets[name[:31]]
                    for cell in ws["B"][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00000E+00"
                    if "Current" in output:
                        for cell in ws["C"][1:]:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "0.00000E+00"

        wb = load_workbook(output_path)

        for ws in wb:
            for col in ws.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in col
                )
                ws.column_dimensions[col[0].column_letter].width = max_length + 5
        wb.save(output_path)
        print(Fore.LIGHTCYAN_EX + f"{board_id} TRACE DOWNLOAD COMPLETE")

def Export_All_Traces(output_folder, phase):
    """Exports every board trace to its own xlsx file by iteratively calling Trace_Getter()"""
    with get_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT board_id
                FROM board_traces
                WHERE phase = %s
                ORDER BY
    CASE
        WHEN board_id ~ '^[0-9]+$'
        THEN board_id::INTEGER
        ELSE NULL
    END,
    board_id
            """, (phase,))

            board_ids = [
                row[0]
                for row in cursor.fetchall()
            ]

    if not board_ids:
        sys.exit("No trace data found.")
    for board_id in board_ids:
        output_path = os.path.join(
            output_folder,
            f"{board_id}_Trace_Data.xlsx"
        )
        Trace_Getter(board_id, phase,output_path)

    print(Fore.GREEN + "ALL TRACE DOWNLOADS COMPLETE")

def Test_Results_Getter(DD, XLSX_NAME,phase) -> None:
    '''Writes out the test data for full database onto one xlsx file'''

    output_file = DD + rf"\{XLSX_NAME}.xlsx"
    with get_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT *
FROM dc_dc_tests
WHERE phase = %s
ORDER BY
    CASE
        WHEN board_id ~ '^[0-9]+$'
        THEN board_id::INTEGER
        ELSE NULL
    END,
    board_id
            """, (phase,))
            rows = cursor.fetchall()

            columns = [
                desc[0]
                for desc in cursor.description
            ]

    df = pd.DataFrame(
        rows,
        columns=columns
    )

    #timestamp reformatting
    df["timestamp"] = pd.to_datetime(df["timestamp"]).dt.strftime("%Y-%m-%d %H:%M")

    debug_remove = ["id","phase"]

    if pc_tests_hide:
        debug_remove = debug_remove + [
            "voltage_dev_warm","voltage_dev_cold","mc_ave_vol","mc_ave_cur","mc_ave_vol_c","mc_ave_cur_c"
        ]
    #same thing but if we want cold deviation only:
    elif cold_only:
        debug_remove = debug_remove + ["voltage_dev_warm","mc_ave_vol","mc_ave_cur"
                                       ]
    if hide_calibration_params:
        debug_remove = debug_remove + ["calibrated_voltage","secondary_calibration"
                                        ]

    #drop the debug cols.
    df = df.drop(columns=debug_remove, errors="ignore")
    df = df.rename(columns={
        "board_id": "BOARD ID",
        "timestamp": "TIMESTAMP",
        "testing_admin": "TEST ADMIN",

        "calibrated_voltage": "CALIBRATED INPUT VOLTAGE (WARM)",
        "initial_voltage": "INITIAL OUTPUT VOLTAGE",
        "initial_current": "INITIAL OUTPUT CURRENT",
        "load_voltage": "LOAD VOLTAGE",
        "load_current": "LOAD CURRENT",

        "initial_start_up": "INITIAL START UP",
        "input_voltage_sweep_w": "INPUT VOLTAGE SWEEP WARM",
        "sweep_min_max_w": "SWEEP MIN/MAX",
        "nominal_load_performance": "NOMINAL LOAD (WARM)",

        "voltage_dev_warm": "WARM VOLTAGE DEVIATION",
        "mc_ave_vol": "WARM POWERCYCLE AVE VOLTAGE",
        "mc_ave_cur": "WARM POWERCYCLE AVE CURRENT",

        "secondary_calibration": "CALIBRATED INPUT VOLTAGE (COLD)",
        "initial_cold_voltage": "INITIAL COLD VOLTAGE",
        "initial_cold_current": "INITIAL COLD CURRENT",
        "load_voltage_c": "LOAD VOLTAGE (COLD)",
        "load_current_c": "LOAD CURRENT (COLD)",

        "input_current_output_voltage": "INITIAL COLD BEHAVIOR",
        "output_step_load": "NOMINAL LOAD (COLD)",
        "input_step_voltage": "INPUT STEP VOLTAGE",
        "input_voltage_sweep_c": "INPUT VOLTAGE SWEEP COLD",
        "sweep_min_max_c": "SWEEP MIN/MAX",
        "cold_start_up": "COLD START UP",

        "voltage_dev_cold": "VOLTAGE DEVIATION (COLD)",
        "mc_ave_vol_c": "AVE VOLTAGE (COLD)",
        "mc_ave_cur_c": "AVE CURRENT (COLD)",
        "shipment": "SHIPMENT",
    })

    df = df.map(convert_scientific_to_float) #convert my prior messy data back
    df.to_excel(output_file, index=False)
    wb = load_workbook(output_file)
    ws = wb.active
    #preserves visible testing categories
    ws.freeze_panes = "A2"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"  # light green
    )
    #sci notation converter
    skip_headers = {
        "BOARD ID",
        "CALIBRATED INPUT VOLTAGE (WARM)",
        "CALIBRATED INPUT VOLTAGE (COLD)"
    }
#a quick reordering since it came out funny
    column_order = [
        "BOARD ID",
        "TIMESTAMP",
        "TEST ADMIN",

        # WARM
        "CALIBRATED INPUT VOLTAGE (WARM)",
        "INITIAL OUTPUT VOLTAGE",
        "INITIAL OUTPUT CURRENT",
        "LOAD VOLTAGE",
        "LOAD CURRENT",
        "INITIAL START UP",
        "INPUT VOLTAGE SWEEP WARM",
        "SWEEP MIN/MAX (WARM)",
        "NOMINAL LOAD (WARM)",
        "WARM VOLTAGE DEVIATION",
        "WARM POWERCYCLE AVE VOLTAGE",
        "WARM POWERCYCLE AVE CURRENT",

        # COLD
        "CALIBRATED INPUT VOLTAGE (COLD)",
        "INITIAL COLD VOLTAGE",
        "INITIAL COLD CURRENT",
        "LOAD VOLTAGE (COLD)",
        "LOAD CURRENT (COLD)",
        "INITIAL COLD BEHAVIOR",
        "NOMINAL LOAD (COLD)",
        "INPUT STEP VOLTAGE",
        "INPUT VOLTAGE SWEEP COLD",
        "SWEEP MIN/MAX (COLD)",
        "COLD START UP",
        "VOLTAGE DEVIATION (COLD)",
        "AVE VOLTAGE (COLD)",
        "AVE CURRENT (COLD)",

        "SHIPMENT",
    ]

    df = df[[col for col in column_order if col in df.columns]]

    for col in ws.iter_cols():
        header = col[0].value

        if header in skip_headers:
            continue

        for cell in col[1:]:  #skip header row
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00000E+00"
#makes it all centered and reformat header row
    thin_black = Side(style="thin", color="D5D5D5")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center")
            cell.border = Border(
                right=thin_black)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    grey_fill = PatternFill(fill_type="solid",fgColor="EBEAEA")
    white_fill = PatternFill(fill_type="solid",fgColor="FFFFFF")
    red_fill = PatternFill(fill_type="solid",fgColor="FFCFCF")
    null_fill = PatternFill(fill_type="solid",fgColor="DFDFAA")

    for row in range(2, ws.max_row + 1):

        fill = grey_fill if row % 2 == 0 else white_fill

        for cell in ws[row]:
            cell.fill = fill

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(
                    max_len,
                    len(str(cell.value))
                )
            if cell.value == "FAIL":
                cell.fill = red_fill
            if cell.value == "NULL" or cell.value == -1.0:
                cell.fill = null_fill

        ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(output_file)

def Shipment_Creator():
    print("\nShipment Creator...\n")
    return

if __name__ == "__main__":
    print(Fore.MAGENTA + "DCDC CONVERTER DATABASE MANAGEMENT TERMINAL\n")
    running = True
    while running:
        TRCHOICE = input(Fore.MAGENTA + ">Download phase SQL data to PC (1)\n"
                                        ">Download phase SQL data to external drive (2)\n"
                                        ">Configure board shipment information (3)\n"
                                        ">Quit data management (4)\n\n"
                                "Choose a readback option (by number): ")

        if TRCHOICE == "1":
            phase = Utilities.SELECT_PHASE()
            folder_path = os.path.join(DD, "DB_IMAGE",phase)
            try:
                os.makedirs(folder_path, exist_ok=True)
                Test_Results_Getter(folder_path, XLSX_NAME,phase)
                print(Fore.GREEN + "TEST DOWNLOAD COMPLETE")
                # now that this saves, lets give it a moment and then write all trace data, this may take some time
                time.sleep(0.2)
                Export_All_Traces(folder_path,phase)

            # just some generic error throws I stole that should be helpful
            except PermissionError:
                print("Error: No permission to write to this drive.")
            except OSError as e:
                print(f"Error creating folder or file: {e}")


        elif TRCHOICE == "2":
            #needs usb or other external mount at D
            drive_letter = "D"
            if os.path.exists(f"{drive_letter.upper()}:\\"):
                print(f"{drive_letter.upper()}: DRIVE FOUND")
                phase = Utilities.SELECT_PHASE()

                folder_path = os.path.join(f"{drive_letter.upper()}:\\", "DB_IMAGE",phase)

                try:
                    #this is the actual writing setp, maybe I should prompt this for every n tests
                    os.makedirs(folder_path, exist_ok=True)
                    Test_Results_Getter(folder_path,XLSX_NAME,phase)
                    print(Fore.GREEN + "TEST DOWNLOAD COMPLETE")
                    #now that this saves, lets give it a moment and then write all trace data, this may take some time
                    time.sleep(0.2)
                    Export_All_Traces(folder_path,phase)

                #just some generic error throws I stole that should be helpful
                except PermissionError:
                    print("Error: No permission to write to this drive.")
                except OSError as e:
                    print(f"Error creating folder or file: {e}")
            else:
                print(f"{drive_letter.upper()}: DRIVE NOT FOUND.")

        elif TRCHOICE == "3":
            #doesn't work yet, I am going to do the traveller manually for now.
            sys.exit()

        elif TRCHOICE == "4":
            sys.exit()

        #super secret admin options
        elif TRCHOICE =="000":
            #admin option to add a new phase name (or rename it)
            ###??? RENAME OPTION DOESNT WORK YET
            phasename = input("Input new phase name   ")
            add_phase(phasename)

        elif TRCHOICE == "-111":
            #admin option to delete a phase
            if input("Delete all empty phases?").lower() == "y":
                delete_empty_phases()

        elif TRCHOICE == "-1":
            #delete specific board
            print()
            print("board deleter")
            phase = SELECT_PHASE()
            Delete_Board_Phase_Entry(input("input board id to delete: "),phase)

        else:
            print(Fore.RED + "INVALID SELECTION")